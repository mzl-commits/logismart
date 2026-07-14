from datetime import timedelta
import io

from django.db import transaction
from django.db.models import Q, Sum
from django.utils import timezone
from django.http import HttpResponse
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import SAFE_METHODS, BasePermission, IsAuthenticated
from rest_framework.response import Response

from .models import Caja, Despacho, DespachoOperacion, MovimientoInventario, PoliticaStock, Producto, ReservaStock, Usuario
from .serializers import MovimientoInventarioSerializer, PoliticaStockSerializer, ProductoSerializer, ReservaStockSerializer


def _available(caja):
    reserved = caja.reservas.filter(estado='activa').aggregate(total=Sum('cantidad'))['total'] or 0
    return max(0, caja.cantidad - reserved)


class IsAdminOrReadOnly(BasePermission):
    def has_permission(self, request, view):
        return request.method in SAFE_METHODS or request.user.is_superuser or request.user.groups.filter(name__in=['Administradores', 'Supervisores']).exists()


def _move(caja, kind, quantity, before, after, request, **extra):
    return MovimientoInventario.objects.create(
        caja=caja, tipo=kind, cantidad=quantity,
        existencia_anterior=before, existencia_posterior=after,
        usuario=request.user if request.user.is_authenticated else None,
        motivo=extra.get('motivo', ''), referencia=extra.get('referencia', ''),
        ubicacion_origen=extra.get('origen'), ubicacion_destino=extra.get('destino'),
    )


class InventoryViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        rows = []
        queryset = Caja.objects.exclude(estado='despachada').prefetch_related('reservas').order_by('producto', 'id')
        search = request.query_params.get('search', '').strip()
        if search:
            queryset = queryset.filter(Q(id__icontains=search) | Q(producto__icontains=search) | Q(codigo_barras__icontains=search) | Q(lote__icontains=search))
        for caja in queryset:
            reserved = sum(r.cantidad for r in caja.reservas.all() if r.estado == 'activa')
            rows.append({'id': caja.id, 'producto': caja.producto, 'codigo_barras': caja.codigo_barras, 'lote': caja.lote,
                         'fecha_vencimiento': caja.fecha_vencimiento, 'fisico': caja.cantidad, 'reservado': reserved,
                         'disponible': max(0, caja.cantidad - reserved), 'estado': caja.estado,
                         'ubicacion': str(caja.id_ubicacion) if caja.id_ubicacion else '',
                         'peso_unitario_kg': float(caja.peso_kg), 'peso_total_kg': float(caja.peso_kg * caja.cantidad)})
        return Response({'items': rows})

    @action(detail=False, methods=['get'])
    def kardex(self, request):
        qs = MovimientoInventario.objects.select_related('caja', 'usuario')
        if request.query_params.get('caja'):
            qs = qs.filter(caja_id=request.query_params['caja'])
        if request.query_params.get('desde'):
            qs = qs.filter(fecha__date__gte=request.query_params['desde'])
        if request.query_params.get('hasta'):
            qs = qs.filter(fecha__date__lte=request.query_params['hasta'])
        return Response(MovimientoInventarioSerializer(qs[:1000], many=True).data)

    @action(detail=False, methods=['get'])
    def exportar_kardex(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        qs = MovimientoInventario.objects.select_related('caja', 'usuario').all()
        if request.query_params.get('desde'):
            qs = qs.filter(fecha__date__gte=request.query_params['desde'])
        if request.query_params.get('hasta'):
            qs = qs.filter(fecha__date__lte=request.query_params['hasta'])
        workbook = Workbook(); sheet = workbook.active; sheet.title = 'Kardex'; sheet.freeze_panes = 'A2'; sheet.sheet_view.showGridLines = False
        sheet.append(['Fecha', 'Caja / SKU', 'Producto', 'Tipo', 'Movimiento', 'Saldo anterior', 'Saldo posterior', 'Usuario', 'Referencia', 'Motivo'])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='E35A16')
        for movement in qs[:10000]:
            sheet.append([timezone.localtime(movement.fecha).replace(tzinfo=None), movement.caja_id, movement.caja.producto,
                          movement.get_tipo_display(), movement.cantidad, movement.existencia_anterior,
                          movement.existencia_posterior, movement.usuario.username if movement.usuario else '',
                          movement.referencia, movement.motivo])
        sheet.column_dimensions['A'].width = 21; sheet.column_dimensions['B'].width = 22; sheet.column_dimensions['C'].width = 28
        for column in 'DEFGHIJ': sheet.column_dimensions[column].width = 18
        for cell in sheet['A'][1:]: cell.number_format = 'dd/mm/yyyy hh:mm'
        output = io.BytesIO(); workbook.save(output)
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="kardex_logismart.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def reservar(self, request):
        quantity = int(request.data.get('cantidad', 0))
        with transaction.atomic():
            caja = Caja.objects.select_for_update().get(pk=request.data.get('caja'))
            if quantity < 1 or quantity > _available(caja):
                return Response({'error': 'Cantidad no disponible para reservar.'}, status=400)
            reservation = ReservaStock.objects.create(caja=caja, cantidad=quantity, destino=request.data.get('destino', ''), creada_por=request.user)
            _move(caja, 'reserva', quantity, caja.cantidad, caja.cantidad, request, motivo='Reserva de stock', referencia=f'RES-{reservation.pk}')
        return Response(ReservaStockSerializer(reservation).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def liberar(self, request):
        with transaction.atomic():
            reservation = ReservaStock.objects.select_for_update().get(pk=request.data.get('reserva'), estado='activa')
            reservation.estado = 'cancelada'; reservation.save(update_fields=['estado', 'actualizada_en'])
            _move(reservation.caja, 'liberacion', reservation.cantidad, reservation.caja.cantidad, reservation.caja.cantidad, request, motivo=request.data.get('motivo', 'Reserva cancelada'), referencia=f'RES-{reservation.pk}')
        return Response({'mensaje': 'Reserva liberada.'})

    @action(detail=False, methods=['post'])
    def despachar(self, request):
        quantity = int(request.data.get('cantidad', 0))
        legacy_user = Usuario.objects.filter(usuario_auth=request.user).first()
        if not legacy_user:
            return Response(
                {'error': 'El usuario autenticado no tiene un perfil logistico asignado.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        with transaction.atomic():
            caja = Caja.objects.select_for_update().get(pk=request.data.get('caja'), estado='almacenada')
            if quantity < 1 or quantity > _available(caja):
                return Response({'error': 'Cantidad no disponible para despacho.'}, status=400)
            before = caja.cantidad
            origin = caja.id_ubicacion
            caja.cantidad -= quantity
            if caja.cantidad == 0:
                caja.estado = 'despachada'; caja.id_ubicacion = None
            caja.save(update_fields=['cantidad', 'estado', 'id_ubicacion'])
            shipment = Despacho.objects.create(id_caja=caja, id_usuario_despacho=legacy_user, cantidad=quantity, destino=request.data.get('destino', 'No especificado'), transporte_placa=request.data.get('transporte_placa', 'N/A'))
            reference = f'DESP-{shipment.pk}'
            _move(caja, 'salida', -quantity, before, caja.cantidad, request, motivo='Despacho parcial' if caja.cantidad else 'Despacho total', referencia=reference, origen=origin)
        return Response({'mensaje': 'Despacho registrado.', 'cantidad_despachada': quantity, 'cantidad_restante': caja.cantidad})

    @action(detail=False, methods=['post'], url_path='despachar_lote')
    def despachar_lote(self, request):
        """Registra todas las salidas de una operación como una sola transacción."""
        idempotency_key = request.headers.get('Idempotency-Key') or request.data.get('idempotency_key')
        if not idempotency_key:
            return Response({'error': 'Falta la clave de idempotencia de la operacion.'}, status=status.HTTP_400_BAD_REQUEST)
        items = request.data.get('items') or []
        if not isinstance(items, list) or not items:
            return Response({'error': 'Debes seleccionar al menos una caja.'}, status=status.HTTP_400_BAD_REQUEST)

        legacy_user = Usuario.objects.filter(usuario_auth=request.user).first()
        if not legacy_user:
            return Response(
                {'error': 'El usuario autenticado no tiene un perfil logistico asignado.', 'code': 'logistic_profile_required'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        previous = DespachoOperacion.objects.filter(clave=idempotency_key, usuario=request.user).first()
        if previous:
            return Response(previous.respuesta, status=status.HTTP_200_OK)

        normalized = []
        seen = set()
        try:
            for item in items:
                caja_id = item.get('caja')
                quantity = int(item.get('cantidad', 0))
                if caja_id in seen:
                    return Response({'error': f'La caja {caja_id} esta repetida en el lote.'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)
                seen.add(caja_id)
                normalized.append((caja_id, quantity))
        except (AttributeError, TypeError, ValueError):
            return Response({'error': 'El formato de las cajas del lote no es valido.'}, status=status.HTTP_422_UNPROCESSABLE_ENTITY)

        locked = []
        with transaction.atomic():
            operation = DespachoOperacion.objects.create(clave=idempotency_key, usuario=request.user)
            for caja_id, quantity in normalized:
                try:
                    caja = Caja.objects.select_for_update().get(pk=caja_id, estado='almacenada')
                except Caja.DoesNotExist:
                    return Response({'error': f'La caja {caja_id} ya no esta disponible para despacho.'}, status=status.HTTP_409_CONFLICT)
                if quantity < 1 or quantity > _available(caja):
                    return Response({'error': f'Cantidad no disponible para la caja {caja_id}.'}, status=status.HTTP_409_CONFLICT)
                locked.append((caja, quantity))

            results = []
            for caja, quantity in locked:
                before = caja.cantidad
                origin = caja.id_ubicacion
                caja.cantidad -= quantity
                if caja.cantidad == 0:
                    caja.estado = 'despachada'
                    caja.id_ubicacion = None
                caja.save(update_fields=['cantidad', 'estado', 'id_ubicacion'])
                shipment = Despacho.objects.create(
                    id_caja=caja,
                    id_usuario_despacho=legacy_user,
                    cantidad=quantity,
                    destino=request.data.get('destino', 'No especificado'),
                    transporte_placa=request.data.get('transporte_placa', 'N/A'),
                )
                _move(caja, 'salida', -quantity, before, caja.cantidad, request,
                      motivo='Despacho parcial' if caja.cantidad else 'Despacho total',
                      referencia=f'DESP-{shipment.pk}', origen=origin)
                if origin and caja.cantidad == 0:
                    origin.estado_ocupacion = False
                    origin.save(update_fields=['estado_ocupacion'])
                results.append({'caja': caja.pk, 'cantidad_despachada': quantity, 'cantidad_restante': caja.cantidad})

            response_data = {'mensaje': 'Despacho por lote registrado.', 'cantidad_despachada': sum(item[1] for item in normalized), 'items': results}
            operation.respuesta = response_data
            operation.save(update_fields=['respuesta'])
        return Response(response_data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def ajustar(self, request):
        new_quantity = int(request.data.get('cantidad', -1))
        if new_quantity < 0 or not request.user.is_superuser:
            return Response({'error': 'Ajuste no autorizado o cantidad inválida.'}, status=403 if not request.user.is_superuser else 400)
        with transaction.atomic():
            caja = Caja.objects.select_for_update().get(pk=request.data.get('caja'))
            before = caja.cantidad; caja.cantidad = new_quantity; caja.save(update_fields=['cantidad'])
            _move(caja, 'ajuste', new_quantity - before, before, new_quantity, request, motivo=request.data.get('motivo', 'Ajuste manual'))
        return Response({'mensaje': 'Existencia ajustada.', 'cantidad': new_quantity})

    @action(detail=False, methods=['get'])
    def alertas(self, request):
        alerts = []
        now = timezone.now()
        for policy in PoliticaStock.objects.filter(activa=True):
            boxes = Caja.objects.filter(producto__iexact=policy.producto).exclude(estado='despachada')
            total = boxes.aggregate(total=Sum('cantidad'))['total'] or 0
            if total <= policy.minimo:
                alerts.append({'tipo': 'stock_bajo', 'producto': policy.producto, 'actual': total, 'limite': policy.minimo})
            if policy.maximo is not None and total >= policy.maximo:
                alerts.append({'tipo': 'stock_alto', 'producto': policy.producto, 'actual': total, 'limite': policy.maximo})
            stale = boxes.filter(hora_llegada__lte=now - timedelta(days=policy.dias_sin_movimiento)).count()
            if stale:
                alerts.append({'tipo': 'inmovilizado', 'producto': policy.producto, 'registros': stale, 'dias': policy.dias_sin_movimiento})
        expiring = Caja.objects.exclude(estado='despachada').filter(fecha_vencimiento__isnull=False, fecha_vencimiento__lte=timezone.localdate() + timedelta(days=30))
        for caja in expiring:
            alerts.append({'tipo': 'vencimiento', 'producto': caja.producto, 'caja': caja.id, 'fecha': caja.fecha_vencimiento})
        return Response(alerts)


class ReservationViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ReservaStock.objects.select_related('caja', 'creada_por')
    serializer_class = ReservaStockSerializer


class StockPolicyViewSet(viewsets.ModelViewSet):
    queryset = PoliticaStock.objects.all()
    serializer_class = PoliticaStockSerializer
    def get_permissions(self):
        return [IsAuthenticated(), IsAdminOrReadOnly()]


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    filterset_fields = ['categoria', 'activo']
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]
