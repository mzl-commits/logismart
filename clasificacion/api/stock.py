"""API REST agrupada por dominio funcional."""

from .common import (
    Caja, Counter, Decimal, HttpResponse, Response, action, io, parse_date,
    status, timezone, viewsets,
)

class StockViewSet(viewsets.ViewSet):
    """Consulta operativa y exportación del inventario registrado."""

    ACTIVE_STATES = ('pendiente', 'clasificada', 'en_transito', 'almacenada')

    def _filtered_queryset(self, request):
        start_raw = request.query_params.get('fecha_desde')
        end_raw = request.query_params.get('fecha_hasta')
        start = parse_date(start_raw) if start_raw else None
        end = parse_date(end_raw) if end_raw else None

        if start_raw and not start:
            raise ValueError('La fecha inicial no es válida.')
        if end_raw and not end:
            raise ValueError('La fecha final no es válida.')
        if start and end and start > end:
            raise ValueError('La fecha inicial no puede ser posterior a la fecha final.')

        queryset = Caja.objects.select_related(
            'id_ubicacion', 'id_medida', 'id_proveedor'
        ).filter(estado__in=self.ACTIVE_STATES).order_by('-hora_llegada', 'producto')

        if start:
            queryset = queryset.filter(hora_llegada__date__gte=start)
        if end:
            queryset = queryset.filter(hora_llegada__date__lte=end)

        search = request.query_params.get('search', '').strip()
        category = request.query_params.get('categoria', '').strip()
        state = request.query_params.get('estado', '').strip()
        if search:
            from django.db.models import Q
            queryset = queryset.filter(Q(id__icontains=search) | Q(producto__icontains=search))
        if category:
            queryset = queryset.filter(categoria=category)
        if state in self.ACTIVE_STATES:
            queryset = queryset.filter(estado=state)

        return queryset, start, end

    @staticmethod
    def _row(caja):
        return {
            'id': caja.id,
            'producto': caja.producto,
            'cantidad': caja.cantidad,
            'categoria': caja.categoria,
            'estado': caja.estado,
            'peso_kg': float(caja.peso_kg),
            'peso_total_kg': float(caja.peso_kg * caja.cantidad),
            'unidad': caja.id_medida.nombre if caja.id_medida else '',
            'proveedor': caja.id_proveedor.nombre_empresa if caja.id_proveedor else '',
            'ubicacion': str(caja.id_ubicacion) if caja.id_ubicacion else 'Sin asignar',
            'fecha_ingreso': timezone.localtime(caja.hora_llegada).isoformat(),
        }

    def list(self, request):
        try:
            queryset, start, end = self._filtered_queryset(request)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        cajas = list(queryset)
        rows = [self._row(caja) for caja in cajas]
        states = Counter(caja.estado for caja in cajas)
        return Response({
            'items': rows,
            'resumen': {
                'referencias': len(cajas),
                'unidades': sum(caja.cantidad for caja in cajas),
                'peso_total_kg': float(sum(
                    (caja.peso_kg * caja.cantidad for caja in cajas), Decimal('0')
                )),
                'almacenadas': states['almacenada'],
                'en_transito': states['en_transito'],
                'pendientes': states['pendiente'] + states['clasificada'],
            },
            'filtros': {
                'fecha_desde': start.isoformat() if start else None,
                'fecha_hasta': end.isoformat() if end else None,
            },
        })

    @action(detail=False, methods=['get'])
    def exportar(self, request):
        try:
            queryset, start, end = self._filtered_queryset(request)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo

        cajas = list(queryset)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Stock'
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = 'A7'

        sheet.merge_cells('A1:K1')
        sheet['A1'] = 'LOGISMART · REPORTE DE STOCK'
        sheet['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill('solid', fgColor='17211B')
        sheet['A1'].alignment = Alignment(vertical='center')
        sheet.row_dimensions[1].height = 32

        period = 'Todo el historial'
        if start or end:
            period = f"{start.strftime('%d/%m/%Y') if start else 'Inicio'} — {end.strftime('%d/%m/%Y') if end else 'Hoy'}"
        sheet['A3'] = 'Periodo de ingreso'
        sheet['B3'] = period
        sheet['D3'] = 'Generado'
        sheet['E3'] = timezone.localtime().replace(tzinfo=None)
        sheet['E3'].number_format = 'dd/mm/yyyy hh:mm'
        sheet['A4'] = 'Referencias'
        sheet['B4'] = len(cajas)
        sheet['D4'] = 'Unidades'
        sheet['E4'] = sum(caja.cantidad for caja in cajas)
        sheet['G4'] = 'Peso total (kg)'
        sheet['H4'] = float(sum((caja.peso_kg * caja.cantidad for caja in cajas), Decimal('0')))
        sheet['H4'].number_format = '#,##0.00'

        headers = ['ID / SKU', 'Producto', 'Categoría', 'Cantidad', 'Unidad', 'Estado',
                   'Peso unit. (kg)', 'Peso total (kg)', 'Ubicación', 'Proveedor', 'Fecha de ingreso']
        sheet.append([])
        sheet.append(headers)
        for caja in cajas:
            sheet.append([
                caja.id, caja.producto, caja.categoria, caja.cantidad,
                caja.id_medida.nombre if caja.id_medida else '',
                caja.get_estado_display(), float(caja.peso_kg),
                float(caja.peso_kg * caja.cantidad),
                str(caja.id_ubicacion) if caja.id_ubicacion else 'Sin asignar',
                caja.id_proveedor.nombre_empresa if caja.id_proveedor else '',
                timezone.localtime(caja.hora_llegada).replace(tzinfo=None),
            ])

        header_fill = PatternFill('solid', fgColor='E35A16')
        thin = Side(style='thin', color='D4D9D6')
        for cell in sheet[6]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(vertical='center')
            cell.border = Border(bottom=thin)
        sheet.row_dimensions[6].height = 26

        for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical='top')
            row[6].number_format = '#,##0.00'
            row[7].number_format = '#,##0.00'
            row[10].number_format = 'dd/mm/yyyy hh:mm'

        if cajas:
            table = ExcelTable(displayName='TablaStock', ref=f'A6:K{sheet.max_row}')
            table.tableStyleInfo = TableStyleInfo(
                name='TableStyleMedium2', showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False,
            )
            sheet.add_table(table)

        widths = {'A': 22, 'B': 28, 'C': 18, 'D': 12, 'E': 13, 'F': 16,
                  'G': 17, 'H': 17, 'I': 24, 'J': 24, 'K': 21}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        summary = workbook.create_sheet('Resumen')
        summary.sheet_view.showGridLines = False
        summary.append(['Categoría', 'Referencias', 'Unidades', 'Peso total (kg)'])
        grouped = {}
        for caja in cajas:
            bucket = grouped.setdefault(caja.categoria or 'otro', [0, 0, Decimal('0')])
            bucket[0] += 1
            bucket[1] += caja.cantidad
            bucket[2] += caja.peso_kg * caja.cantidad
        for category, values in sorted(grouped.items()):
            summary.append([category, values[0], values[1], float(values[2])])
        for cell in summary[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
        summary.freeze_panes = 'A2'
        summary.column_dimensions['A'].width = 24
        for column in ('B', 'C', 'D'):
            summary.column_dimensions[column].width = 18
        for cell in summary['D'][1:]:
            cell.number_format = '#,##0.00'

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = f"stock_{(start or timezone.localdate()).strftime('%Y%m%d')}_{(end or timezone.localdate()).strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


